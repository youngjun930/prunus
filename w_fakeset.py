import json
import os
import openpyxl
from _collections import OrderedDict

os.chdir("D:/프로젝트 데이터/YW모바일/돌아온페이크셋 - 복사본")                       # 저장할 경로 입력
file_list = os.listdir()

wb = openpyxl.load_workbook("D:/프로젝트 데이터/YW모바일/페이크셋.xlsx")
ws = wb.get_sheet_by_name("페이크셋")

for i in range(90):

    file_data = OrderedDict()

    file_data["content"] = "{}".format(ws["B{}".format(i+2)].value)  # 사진url 혹은 텍스트 입력
    file_data["keyword"] = ""  # 키워드 텍스트입력, 없으면 ""
    file_data["missions"] = [
        {
            "inputId": 392,  # inputid를 입력
            "textAnswer": "",  # 인풋이 텍스트인풋 or 라인그리드  string으로 입력("") 아닌경우 None
            "optionIds": [ws["E{}".format(i+2)].value],  # 드롭다운, 라디오, 체크박스일 시 정답 optionid 입력, 아닌경우 []
            "questionAndAnswers": [] # 그리드라디오, 그리드체크박스일때 {"questionId": QuestionOptionId, "answerIds": List Of Answer Option Id} 딕셔너리 입력, 아닌경우 []
        },
        {
            "inputId": 393,
            "textAnswer": "",
            "optionIds": [ws["F{}".format(i+2)].value],
            "questionAndAnswers": []
        }
    ]

    print(json.dumps(file_data, ensure_ascii=False, indent="\t"))

    with open("{}.json".format(ws["A{}".format(i+2)].value), "w", encoding="utf-8") as make_file:  # 파일명 입력
        json.dump(file_data, make_file, ensure_ascii=False, indent="\t")

